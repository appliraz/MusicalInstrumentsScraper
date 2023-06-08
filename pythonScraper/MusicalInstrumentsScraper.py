import AviGilScraper, DiezScraper, HalilitScraper, KleyZemerScraper, MusicCenterScraper, WildScraper
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


avigil = ["avigil.co.il", AviGilScraper,"אבי גיל"]
diez = ["diez.co.il", DiezScraper, "דיאז"]
halilit = ["halilit.com", HalilitScraper, "חלילית"]
kley_zemer = ["kley-zemer.co.il", KleyZemerScraper, "כלי זמר"]
music_center = ["music-center.co.il", MusicCenterScraper, "מיוזיק סנטר"]
wildguitars = ["wildguitars.co.il", WildScraper, "ווילד גיטרס"]

"""
avigil = [{"website": "avigil.co.il", "scraper": AviGilScraper, "name": "avi-gil"}]
diez = [{"website": "diez.co.il", "scraper": DiezScraper, "name": "diez"}]
halilit = [{"website": "halilit.com", "scraper": HalilitScraper, "name": "halilit"}]
kley_zemer = [{"website": "kley-zemer.co.il","scraper": KleyZemerScraper, "name":"kley-zemer"}]
music_center = [{"website": "music-center.co.il", "scraper": MusicCenterScraper, "name": "music-center"}]
wildguitars = [{"website": "wildguitars.co.il", "scraper": WildScraper, "name": "wildguitars"}]
"""

websites = [avigil, diez, halilit, kley_zemer, music_center, wildguitars]

def getAllowedWebsites():
    allowed = []
    for website in websites:
        allowed.append((website[0], website[2]))
    return allowed

def getScraper(url):
    print(f"finding the relevant scraper for url {url}")
    for website, scraper, name in websites:
        if website in url:
             return (scraper, name)
    return Exception

def getScrapedData(url, scraper):
    print('started the scraping')
    try:
        data = scraper.scrap(url)
    except Exception as e:
        print("failed in scraping data, sending empty array")
        print(e)
        data = ['no', 'scraping', 'here']
    return data

def reformatSheet(sheet):
    # make the first row bold
    print(f'reformating sheet')
    row =sheet[1]
    for cell in row:
        cell.font = Font(bold=True)
    
    # make the price column in number format
    max_row = sheet.max_row
    max_col = sheet.max_column

    for i in range(1, max_row+1):
        cell_index = "C" + str(i)
        try:
            sheet[cell_index].value = float(sheet[cell_index].value)
        except Exception as e:
            print(e)
            continue
    
    # adjust the width of the columns and aligment
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 40

    for row in sheet.iter_rows(min_row=1, max_col=max_col, max_row=max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def writeDataToWorkbook(data, name, wb: Workbook):
    print(f'writing data to sheet {name}')
    ws = wb.create_sheet(name)
    #add headers
    headers = ['קישור', 'שם המוצר', 'מחיר', 'מלאי']
    ws.append(headers)

    #add data
    for row in data:
        ws.append(row)

    #reformate the sheet
    reformatSheet(ws)

def deleteUnnessacerySheets(wb: Workbook):
    sheets_to_delete = ['Sheet1', 'Sheet']
    for sheet in sheets_to_delete:
        if sheet in wb.sheetnames:
            del wb[sheet]


def scrapToExcel(urls, filename="TheMusicalInstrumentsScraper"):
    print('start scraping')
    sheet_no = 1
    wb = Workbook()
    print("urls = ")
    print(urls) 
    if not urls:
        raise TypeError("no urls received")
    for url in urls:
        print("url: " + url)
        try:
            scraper, name = getScraper(url)
            data = getScrapedData(url, scraper)
            name_of_sheet = name + " " + str(sheet_no)
            writeDataToWorkbook(data, name_of_sheet , wb)
        except Exception as e:
            print(e)
            continue
        sheet_no += 1
    deleteUnnessacerySheets(wb)
    wb.save(filename)

#main(["https://www.halilit.com/23604-Studio-Monitors", "https://diez.co.il/product-category/%d7%a1%d7%90%d7%95%d7%a0%d7%93-%d7%95%d7%94%d7%92%d7%91%d7%a8%d7%94/%d7%90%d7%95%d7%9c%d7%a4%d7%9f/%d7%9e%d7%95%d7%a0%d7%99%d7%98%d7%95%d7%a8%d7%99%d7%9d-%d7%90%d7%95%d7%9c%d7%a4%d7%a0%d7%99%d7%99%d7%9d/"])